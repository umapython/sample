from django.shortcuts import render,redirect
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
from .models import Topic,Answertable
# Create your views here.
from django.http import HttpResponse,request

def index(request):
    #return HttpResponse('<h1>Home Page</h1>')
    return render(request,'wwcodeapp/index.html')

def details(request):
    return render(request,'wwcodeapp/details.html')

def datas(request):
    if request.method == "GET":
        return render(request,'wwcodeapp/datas.html')
    else:
        excel_file = request.FILES["excel_file"]
        
        #print(excel_file)
        currentdir = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(currentdir,"trial.xlsx")
        #print(path)
        xlsx_file = Path(currentdir, 'trial.xlsx')
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        excel_data = {}
        i=1
        row_count = sheet.max_row
        column_count = sheet.max_column
        for i in range(2, row_count + 1):
            new_list = []
            for j in range(1, column_count + 1):
                data = sheet.cell(row=i, column=j).value
                #print(data, end='   ')
                new_list.append(data)
            #print('\n')
            excel_data[i]=new_list

        # iterating over the rows and
        # getting value from each cell in row
        #print(excel_data)
        #print("====")
        return render(request,'wwcodeapp/datas.html',{"excel_data":excel_data})

def cards_save(request,id):
    
    return redirect('cards')

def cards(request,**kwargs): 
    for key, value in kwargs.items (): 
        print ("%s == %s" %(key, value))
        if value:
            temp=value
        temp="id"+str(value)
        print(temp)
        #keys=form.cleaned_data[key]
        #print(keys)
    #if request.method == "GET":
        answer=request.GET[temp]
        print(answer)
        reg=Answertable(Answer = answer)
        reg.save()
    all_objects= Topic.objects.all()   
    qbank = [{ blog.pk: blog.Question} for blog in all_objects]
    #print(qbank)
  
    return render(request,'wwcodeapp/cards.html',{"qbank":qbank})